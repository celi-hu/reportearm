import math
from zoneinfo import ZoneInfo
import pandas as pd
from typing import Any, Dict, List, Optional
import time
import requests
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Tuple, Optional
import pandas as pd
import re
from openpyxl import load_workbook


import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill


TZ_AR = ZoneInfo("America/Argentina/Buenos_Aires")

# =========================
# CONSTANTES
# =========================

HORAS_A_VACIAR_DIA_ANTERIOR = [
    "Horas Trabajadas",
    "HORAS_REGULAR",
    "HORAS_EXTRA",
    "HORAS_EXTRA AL 50",
    "HORAS_EXTRA AL 100",
    "HORAS_NOCTURNA",
    "HORAS_FRANCO",
    "HORAS_FERIADO",
    "HORAS_FERIADO NOCTURNA",
    "HORAS_FRANCO NOCTURNA",
    "HORAS_NOCTURNA 2",
    "HORAS_EXTRA SABADO",
    "HORAS_EXTRA DOMINGO",
]

INCIDENCES_MAP = {
    "ABSENT": "Ausencia sin aviso",
    "LATE": "Tardanza",
    "UNDERWORKED": "Trabajo insuficiente",
    "LOCATION_INCIDENCE": "Fuera de ubicación"
}

WEEKDAY_ES_MAP = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}

# =========================
# FECHAS / HORAS
# =========================

def iso_to_dt(value, tz=TZ_AR):
    if not value:
        return pd.NaT
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(tz)
    except Exception:
        return pd.NaT


def floor_minute(dt):
    if pd.isna(dt):
        return dt
    return dt.replace(second=0, microsecond=0)


def weekday_es(date_str: str) -> str:
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return WEEKDAY_ES_MAP[d.weekday()]
    except Exception:
        return ""


def fmt_range(start, end):
    if pd.isna(start) or pd.isna(end):
        return ""
    return f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"


def calc_delta_hours(real, sched, tolerance_seconds=0):
    if pd.isna(real) or pd.isna(sched):
        return 0.0
    delta = (real - sched).total_seconds() - tolerance_seconds
    return round(delta / 3600, 2) if delta > 0 else 0.0

def calc_early_arrival_hours(real, sched):
    if pd.isna(real) or pd.isna(sched):
        return 0.0

    delta = (sched - real).total_seconds()
    return round(delta / 3600, 2) if delta > 0 else 0.0


# =========================
# CATEGORÍAS / HORAS
# =========================

def split_categorized_hours(categorized_hours, categorias_validas):
    """
    Devuelve dict:
    HORAS_<CATEGORIA> = horas
    """
    out = {}
    for c in categorias_validas:
        out[f"HORAS_{c}"] = 0.0

    for ch in categorized_hours or []:
        name = (ch.get("category", {}).get("name") or "").upper().strip()
        if name in categorias_validas:
            out[f"HORAS_{name}"] += float(ch.get("hours") or 0)

    return {k: round(v, 2) for k, v in out.items()}


# =========================
# INCIDENCIAS
# =========================
def build_observaciones(it: dict) -> str:
    obs = []

    # =========================
    # Feriados
    # =========================
    holidays = it.get("holidays") or []
    if isinstance(holidays, list) and holidays:
        names = []
        for h in holidays:
            if isinstance(h, dict):
                n = (h.get("name") or "").strip()
                if n:
                    names.append(n)

        if names:
            # dedupe simple (orden no importa tanto en feriados)
            obs.append("Feriado: " + " | ".join(sorted(set(names))))

    # =========================
    # Incidencias
    # - puede venir como ["ABSENT", ...]
    # - o como [{"name": "..."}]
    # =========================
    incidences = it.get("incidences") or []
    if isinstance(incidences, list) and incidences:
        names = []
        for inc in incidences:
            if isinstance(inc, str):
                n = inc.strip()
                if n:
                    key = n.upper().strip()
                    label = INCIDENCES_MAP.get(key, n)  # 👈 traducción
                    names.append(label)
            elif isinstance(inc, dict):
                n = (inc.get("name") or inc.get("type") or inc.get("code") or "").strip()
                if n:
                    names.append(n)

        if names:
            # dedupe manteniendo orden
            seen = set()
            names_unique = []
            for n in names:
                if n not in seen:
                    seen.add(n)
                    names_unique.append(n)

            obs.append("Incidencia: " + " | ".join(names_unique))

    # =========================
    # Time off requests (Licencias)
    # =========================
    tors = it.get("timeOffRequests") or []
    if isinstance(tors, list) and tors:
        names = []
        for tor in tors:
            if isinstance(tor, dict):
                n = (tor.get("name") or "").strip()
                if n:
                    names.append(n)

        if names:
            # dedupe manteniendo orden
            seen = set()
            names_unique = []
            for n in names:
                if n not in seen:
                    seen.add(n)
                    names_unique.append(n)

            obs.append("Licencia: " + " | ".join(names_unique))

    return " | ".join(obs)


def clasificar_empleado_por_scheduled_max(df, col_sched="SCHEDULED_HOURS"):
    """
    FULL-TIME  : max scheduled >= 8
    PART-TIME  : max scheduled < 8
    """
    res = {}
    grouped = df.groupby("ID")[col_sched].max()

    for emp, max_h in grouped.items():
        try:
            max_h = float(max_h or 0)
        except Exception:
            max_h = 0

        if max_h >= 8:
            res[emp] = "FULL-TIME"
        else:
            res[emp] = "PART-TIME"

    return res


# =========================
# NOCTURNIDAD
# =========================

def nocturnidad_es_100(row):
    """
    Regla:
    - Domingo
    - Feriado
    - No laborable
    - Sábado (fallback conservador)
    """
    weekday = row.get("_weekday_api", "")
    if weekday in ("SUNDAY",):
        return True

    if row.get("_hasHoliday_api", False):
        return True

    if row.get("_isWorkday_api") is False:
        return True

    if weekday == "SATURDAY":
        return True

    return False


# =========================
# STRINGS / NOMBRES
# =========================

def split_apellido_nombre(value):
    if not value or "," not in value:
        return "", ""
    apellido, nombre = value.split(",", 1)
    return apellido.strip(), nombre.strip()


# =========================
# EXPORTACIÓN EXCEL
# =========================

def horas_para_excel(value, usar_decimal=True):
    """
    - 0 => celda vacía
    - decimal => float
    - hh:mm => fracción de día
    """
    try:
        v = float(value)
    except Exception:
        return ""

    if v == 0:
        return ""

    if usar_decimal:
        return round(v, 2)

    return v / 24.0



def export_detalle_diario_excel(
    df_export: pd.DataFrame,
    out: str,
    START_DATE: str,
    END_DATE: str,
    generated_at: str,
    EXPORTAR_DECIMAL: bool,
    COLS_HORAS_DETALLE: list,
):
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        workbook = writer.book

        fmt_title = workbook.add_format({"bold": True, "font_size": 14})
        fmt_sub   = workbook.add_format({"font_size": 11})
        fmt_wrap  = workbook.add_format({"text_wrap": True})
        fmt_hhmm  = workbook.add_format({"num_format": "[h]:mm"}) if not EXPORTAR_DECIMAL else None
        fmt_dec   = workbook.add_format({"num_format": "0.00"})   if EXPORTAR_DECIMAL else None

        # ----- Detalle diario -----
        startrow = 4

        # Ordenar
        df_export = df_export.sort_values(by=["ID", "Fecha"], ascending=[True, True]).reset_index(drop=True)

        # Escribir dataframe
        df_export.to_excel(writer, index=False, sheet_name="Detalle diario", startrow=startrow)
        ws1 = writer.sheets["Detalle diario"]

        # Encabezado
        ws1.write("A1", "DETALLE DIARIO DE ASISTENCIA", fmt_title)
        ws1.write("A2", f"Período: {START_DATE} al {END_DATE}", fmt_sub)
        ws1.write("A3", f"Generado: {generated_at}", fmt_sub)

        # Anchos + formatos
        for idx, col in enumerate(df_export.columns):
            if col == "Observaciones":
                ws1.set_column(idx, idx, 45, fmt_wrap)
            elif col in COLS_HORAS_DETALLE:
                ws1.set_column(idx, idx, 22, fmt_dec if EXPORTAR_DECIMAL else fmt_hhmm)
            else:
                ws1.set_column(idx, idx, 26)

    print("✅ Excel generado:", out)


def aplicar_regla_extra_50(horas_extra):
    """
    Regla:
    - <= 0.5  -> 0
    - > 0.5   -> baja a escalones de 0.5
      0.6..1.0 -> 0.5
      1.1..1.5 -> 1.0
      1.6..2.0 -> 1.5
      etc.
    """
    try:
        h = float(horas_extra)
    except Exception:
        return 0.0

    if h <= 0.5:
        return 0.0

    # ejemplo: 1.1 -> int((1.1 - 0.5) / 0.5) = 1  -> 0.5
    escalones = int((h - 0.5) / 0.5)
    return round(escalones * 0.5, 2)


def decimal_to_hhmm(x):
    if pd.isna(x) or x == "":
        return ""
    minutes = int(round(float(x) * 60))
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"



def flags_incidencias_y_eventos(
    entries=None,
    slots=None,
    incidences=None,
    time_off_requests=None,
    holidays=None
) -> dict:
    """
    Devuelve flags 'Si' / 'No' para columnas:
    - Ausencia
    - Tardanza
    - Retiro anticipado
    - Trabajo Insuficiente
    - Es Feriado
    - Licencia
    """

    incidences = incidences or []
    time_off_requests = time_off_requests or []
    holidays = holidays or []

    # Normalizar incidencias a set de claves UPPER
    inc_keys = set()

    for inc in incidences:
        if isinstance(inc, str):
            key = inc.strip().upper()
        elif isinstance(inc, dict):
            key = (inc.get("name") or inc.get("type") or inc.get("code") or "").strip().upper()
        else:
            continue

        if key:
            inc_keys.add(key)

    return {
        "Ausencia": "Si" if "ABSENT" in inc_keys else "No",
        "Tardanza": "Si" if "LATE" in inc_keys else "No",
        "Trabajo Insuficiente": "Si" if "UNDERWORKED" in inc_keys else "No",
        "Es Feriado": "Si" if bool(holidays) else "No",
        "Licencia": "Si" if bool(time_off_requests) else "No",
    }

def pintar_flags_si_no(
    path_xlsx: str,
    sheet_name: str = "Detalle diario",
    cols_flag=None,
):
    if cols_flag is None:
        cols_flag = [
            "Ausencia",
            "Tardanza -",
            "Trabajo Insuficiente",
            "Es Feriado",
            "Licencia",
            "Cruce de día",
            "Ajuste cruce→feriado",
        ]

    wb = load_workbook(path_xlsx)
    ws = wb[sheet_name]

    # Buscar fila de headers (donde están los nombres exactos)
    header_row = None
    header_map = {}  # nombre_col -> idx_col (1-based)

    for r in range(1, 31):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row_vals_str = [("" if v is None else str(v).strip()) for v in row_vals]

        # Si al menos 2 de las cols_flag aparecen, asumimos que es el header
        hits = sum(1 for name in cols_flag if name in row_vals_str)
        if hits >= 2:
            header_row = r
            for c, v in enumerate(row_vals_str, start=1):
                if v:
                    header_map[v] = c
            break

    if header_row is None:
        wb.close()
        raise ValueError(f"No encontré la fila de headers en '{sheet_name}' (busqué hasta la fila 30).")

    # Resolver columnas a pintar (solo las que existen)
    cols_idx = []
    for name in cols_flag:
        if name in header_map:
            cols_idx.append((name, header_map[name]))

    # Colores (claros)
    fill_green = PatternFill("solid", fgColor="C6EFCE")  # verde claro
    fill_red   = PatternFill("solid", fgColor="FFC7CE")  # rojo claro

    def norm(v):
        if v is None:
            return ""
        return str(v).strip().lower()

    # Pintar celdas
    for r in range(header_row + 1, ws.max_row + 1):
        for name, c in cols_idx:
            v = norm(ws.cell(r, c).value)
            if v in ("si", "sí"):
                ws.cell(r, c).fill = fill_green
            elif v == "no":
                ws.cell(r, c).fill = fill_red
            # else: no tocar (vacío u otro valor)

    wb.save(path_xlsx)
    wb.close()


def redondear_extra_media_hora(x):
    # x en horas decimales (float)
    v = pd.to_numeric(x, errors="coerce")
    if pd.isna(v) or v < 0.5:
        return 0.0
    return math.floor(v / 0.5) * 0.5




COLS_SUM = [
    "Horas Trabajadas",
    "HORAS_FRANCO",
    "HORAS_FERIADO",
    "HORAS_EXTRA AL 50",
    "HORAS_EXTRA AL 100",
    "TARDANZA",
]

def _read_detalle_diario_from_export(path_xlsx: str) -> pd.DataFrame:
    """
    Tu 'Detalle diario' tiene 3 filas arriba (Período/Generado/vacío) y luego el header real.
    Esto lo lee robusto y devuelve el DataFrame con columnas correctas.
    """
    raw = pd.read_excel(path_xlsx, sheet_name="Detalle diario", skiprows=3)
    cols = raw.iloc[0].tolist()
    df = raw.iloc[1:].reset_index(drop=True)
    df.columns = cols
    return df

def _build_resumen(df_det: pd.DataFrame, turno_key: str) -> pd.DataFrame:
    df = df_det.copy()

    # numéricos
    for c in COLS_SUM:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    # normalizar turno (minúsculas + trim)
    df["Turno_norm"] = df["Turno"].astype(str).str.strip().str.lower()

    dft = df[df["Turno_norm"] == turno_key].copy()
    if dft.empty:
        return pd.DataFrame(columns=["ID", "Apellido, Nombre", "Turno"] + COLS_SUM)

    res = (
        dft.groupby("ID", as_index=False)
           .agg({**{"Apellido, Nombre": "first", "Turno": "first"},
                 **{c: "sum" for c in COLS_SUM}})
    )

    res = res[["ID", "Apellido, Nombre", "Turno"] + COLS_SUM]
    for c in COLS_SUM:
        res[c] = res[c].round(2)

    return res.sort_values(["Apellido, Nombre", "ID"], kind="mergesort").reset_index(drop=True)

def _write_df_to_sheet(wb, sheet_name: str, df_out: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # escribir DF
    for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c in range(1, len(row) + 1):
                cell = ws.cell(r_idx, c)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill("solid", fgColor="E7EEF9")  # header suave

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # widths básicos
    widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[col].width = min(max(10, w + 2), 40)

    # formato numérico
    for col_idx, col_name in enumerate(df_out.columns, start=1):
        if col_name in COLS_SUM:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_idx).number_format = "0.00"

def agregar_resumen_turnos(path_xlsx: str):
    df_det = _read_detalle_diario_from_export(path_xlsx)

    resumen_tarde  = _build_resumen(df_det, "tarde")
    resumen_manana = _build_resumen(df_det, "mañana")

    wb = load_workbook(path_xlsx)
    _write_df_to_sheet(wb, "Resumen Tarde", resumen_tarde)
    _write_df_to_sheet(wb, "Resumen Mañana", resumen_manana)
    wb.save(path_xlsx)
    wb.close()


def worked_hours_from_entries(rs, re) -> float:
    """Horas reales trabajadas según fichadas (rs/re). Devuelve 0 si falta algo."""
    if rs is None or pd.isna(rs) or re is None or pd.isna(re):
        return 0.0
    try:
        delta = (re - rs).total_seconds() / 3600.0
        return round(max(0.0, delta), 2)
    except Exception:
        return 0.0
    


def restar_llegada_anticipada(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    eps = 1e-9

    def num(col, default=0.0):
        if col not in out.columns:
            out[col] = default
        return pd.to_numeric(out[col], errors="coerce").fillna(default).astype(float)

    llegada = num("LLEGADA_ANTICIPADA")
    worked  = num("Horas Trabajadas")
    planif  = num("Horas planificadas")
    feriado = num("HORAS_FERIADO")
    franco  = num("HORAS_FRANCO")
    ex50    = num("HORAS_EXTRA AL 50")
    ex100   = num("HORAS_EXTRA AL 100")
    extra_t = num("HORAS_EXTRA")

    is_franco  = franco > eps
    is_feriado = (~is_franco) & (feriado > eps)
    is_normal  = (~is_franco) & (~is_feriado)

    # ── FRANCO: no se toca nada ──
    feriado_new = feriado.copy()
    worked_new  = worked.copy()
    ex50_new    = ex50.copy()
    ex100_new   = ex100.copy()
    extra_new   = extra_t.copy()

    # ── FERIADO: restar llegada de feriado y trabajadas ──
    #feriado_new = np.where(is_feriado, np.maximum(feriado - llegada, 0.0), feriado_new)
    #worked_new  = np.where(is_feriado, np.maximum(worked  - llegada, 0.0), worked_new)
    pass

    # ── NORMAL ──
    horas_netas = np.where(is_normal, extra_t - llegada, 0.0)

    # Netas < 0: llegada anticipada > extra → extra a 0, se descuenta de trabajadas
    mask_neg = is_normal & (horas_netas < -eps) & (worked > eps)

    ex50_new  = np.where(mask_neg, 0.0, ex50_new)
    ex100_new = np.where(mask_neg, 0.0, ex100_new)
    extra_new = np.where(mask_neg, 0.0, extra_new)
    worked_new = np.where(mask_neg, np.maximum(planif + horas_netas, 0.0), worked_new)

    # Netas >= 0: llegada anticipada <= extra → restar de extra Y de trabajadas
    mask_pos = is_normal & (horas_netas >= -eps) & (worked > eps)
    has_100 = mask_pos & (ex100 > eps)
    has_50  = mask_pos & (~has_100) & (ex50 > eps)

    ex100_new  = np.where(has_100, np.maximum(ex100  - llegada, 0.0), ex100_new)
    ex50_new   = np.where(has_50,  np.maximum(ex50   - llegada, 0.0), ex50_new)
    extra_new  = np.where(mask_pos, np.maximum(extra_t - llegada, 0.0), extra_new)
    worked_new = np.where(mask_pos, np.maximum(worked  - llegada, 0.0), worked_new)

    out["HORAS_FERIADO"]      = np.round(feriado_new, 2)
    out["Horas Trabajadas"]   = np.round(worked_new,  2)
    out["HORAS_EXTRA AL 50"]  = np.round(ex50_new,    2)
    out["HORAS_EXTRA AL 100"] = np.round(ex100_new,   2)
    out["HORAS_EXTRA"]        = np.round(extra_new,   2)
    out["Horas_Netas"]        = np.round(horas_netas, 2)

    return out

def agregar_resumen_general(path_xlsx: str):
    df_det = _read_detalle_diario_from_export(path_xlsx)

    df = df_det.copy()

    for c in COLS_SUM:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    res = (
        df.groupby("ID", as_index=False)
          .agg({**{"Apellido, Nombre": "first", "Turno": "first"},
                **{c: "sum" for c in COLS_SUM}})
    )

    res = res[["ID", "Apellido, Nombre", "Turno"] + COLS_SUM]
    for c in COLS_SUM:
        res[c] = res[c].round(2)

    res = res.sort_values(["Apellido, Nombre", "ID"], kind="mergesort").reset_index(drop=True)

    wb = load_workbook(path_xlsx)
    _write_df_to_sheet(wb, "Resumen General", res)
    wb.save(path_xlsx)
    wb.close()

def aplicar_ajuste_cruce_a_feriado(df_export: pd.DataFrame) -> pd.DataFrame:
    """
    Corrige HORAS_FERIADO para turnos nocturnos que cruzan días, según estas reglas:

    Regla A: inicio NO feriado → fin FERIADO
        → HORAS_FERIADO se mueve al día siguiente. El día de inicio: HORAS_FERIADO = 0.

    Regla B: inicio FERIADO → fin FERIADO
        → HORAS_FERIADO del día siguiente se acumula en el día de inicio. 
          El día siguiente: HORAS_FERIADO = 0.
          Solo aplica si el día de inicio NO fue ya destino de Regla A.

    Regla C: inicio FERIADO → fin NO feriado
        → HORAS_FERIADO = 0 en el día de inicio.
          Solo aplica si el día de inicio NO fue ya destino de Regla A.

    NO toca Horas Trabajadas ni ninguna otra columna.
    """
    df = df_export.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.sort_values(["ID", "Fecha"]).reset_index(drop=True)

    def _safe_float(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, pd.Timestamp) or v is pd.NaT:
            return 0.0
        try:
            if pd.isna(v):
                return 0.0
        except (TypeError, ValueError):
            pass
        try:
            val = float(v)
            if val > 200 or val < 0:
                return 0.0
            return val
        except Exception:
            return 0.0

    def _es_feriado(row) -> bool:
        obs = str(row.get("Observaciones") or "").lower()
        return "feriado" in obs or _safe_float(row.get("HORAS_FERIADO", 0)) > 0

    def _cruce_de_dia(row) -> bool:
        return str(row.get("Cruce de día") or "").strip().lower() == "si"

    if "HORAS_FERIADO" not in df.columns:
        df["HORAS_FERIADO"] = 0.0
    else:
        df["HORAS_FERIADO"] = df["HORAS_FERIADO"].apply(_safe_float)

    for uid, idxs in df.groupby("ID").groups.items():
        idxs = list(idxs)
        procesados = set()

        for j in range(len(idxs) - 1):
            i_cur = idxs[j]
            i_nxt = idxs[j + 1]

            row_cur = df.loc[i_cur]
            row_nxt = df.loc[i_nxt]

            if not _cruce_de_dia(row_cur):
                continue

            cur_es_feriado = _es_feriado(row_cur)
            nxt_es_feriado = _es_feriado(row_nxt)

            cur_feriado = _safe_float(row_cur.get("HORAS_FERIADO", 0))
            nxt_feriado = _safe_float(row_nxt.get("HORAS_FERIADO", 0))
            cur_worked  = _safe_float(row_cur.get("Horas Trabajadas", 0))
            nxt_worked  = _safe_float(row_nxt.get("Horas Trabajadas", 0))

            # ── REGLA A: NO feriado → FERIADO ──
            if not cur_es_feriado and nxt_es_feriado:
                df.at[i_nxt, "HORAS_FERIADO"] = round(cur_worked, 2)
                df.at[i_cur, "HORAS_FERIADO"] = 0.0
                df.at[i_cur, "HORAS_EXTRA AL 100"] = 0.0
                df.at[i_cur, "HORAS_EXTRA AL 50"]  = 0.0
                df.at[i_cur, "_regla_a_aplicada"]  = True  # ← flag
                procesados.add(i_nxt)

                obs_cur = str(df.at[i_cur, "Observaciones"] or "").strip()
                df.at[i_cur, "Observaciones"] = (
                    obs_cur + f" | Regla A: horas feriado pasadas al {row_nxt['Fecha'].date()}"
                ).strip(" |")
                obs_nxt = str(df.at[i_nxt, "Observaciones"] or "").strip()


            # ── REGLA B: FERIADO → FERIADO ──
            elif cur_es_feriado and nxt_es_feriado and i_cur not in procesados:
                df.at[i_cur, "HORAS_FERIADO"] = round(cur_feriado + nxt_worked, 2)
                df.at[i_nxt, "HORAS_FERIADO"] = 0.0
                procesados.add(i_cur)

                obs_nxt = str(df.at[i_nxt, "Observaciones"] or "").strip()
                df.at[i_nxt, "Observaciones"] = (
                    obs_nxt + f" | Regla B: horas consolidadas en {row_cur['Fecha'].date()}"
                ).strip(" |")


            # ── REGLA C: FERIADO → NO feriado ──
            elif cur_es_feriado and not nxt_es_feriado and i_cur not in procesados:
                cur_w = _safe_float(row_cur.get("Horas Trabajadas", 0))
                franco_cur = _safe_float(row_cur.get("HORAS_FRANCO", 0))

                # Limpiar categorías del día de inicio
                df.at[i_cur, "HORAS_FERIADO"] = 0.0
                df.at[i_cur, "HORAS_EXTRA AL 100"] = 0.0
                df.at[i_cur, "HORAS_EXTRA AL 50"] = 0.0
                df.at[i_cur, "Horas Trabajadas"] = 0.0

                if cur_w > 0:
                    if franco_cur > 0:
                        # Feriado coincide con franco del empleado
                        # → horas en inicio como FRANCO, AL100 lo asigna aplicar_prioridades_horas_extra
                        # → día siguiente no se toca
                        df.at[i_cur, "HORAS_FRANCO"] = franco_cur
                        df.at[i_cur, "Horas Trabajadas"] = cur_w
                        obs_cur = str(df.at[i_cur, "Observaciones"] or "").strip()
                        df.at[i_cur, "Observaciones"] = (
                                obs_cur + " | Regla C: franco coincide con feriado → AL 100"
                        ).strip(" |")
                    else:
                        # Feriado no coincide con franco → solo Horas Trabajadas en inicio
                        # → día siguiente se limpia
                        df.at[i_cur, "Horas Trabajadas"] = cur_w
                        df.at[i_nxt, "Horas Trabajadas"] = 0.0
                        df.at[i_nxt, "HORAS_EXTRA AL 100"] = 0.0
                        df.at[i_nxt, "HORAS_EXTRA AL 50"] = 0.0
                        df.at[i_nxt, "HORAS_FRANCO"] = 0.0
                        obs_cur = str(df.at[i_cur, "Observaciones"] or "").strip()
                        df.at[i_cur, "Observaciones"] = (
                                obs_cur + " | Regla C: horas trabajadas en feriado sin categoría extra"
                        ).strip(" |")
    df["HORAS_FERIADO"] = pd.to_numeric(df["HORAS_FERIADO"], errors="coerce").fillna(0.0).round(2)

    return df




def calcular_cruce_dia(real_start, real_end) -> str:
    if real_start is None or pd.isna(real_start) or real_end is None or pd.isna(real_end):
        return "No"
    try:
        return "Si" if real_start.date() != real_end.date() else "No"
    except Exception:
        return "No"
    

# ================= CATEGORÍAS =================
def split_categorized_hours_basic(categorized_hours, categorias_validas):
    valid_upper = {c.upper(): c for c in categorias_validas}
    out = {f"HORAS_{c}": 0.0 for c in categorias_validas}

    for ch in categorized_hours or []:
        name = (ch.get("category", {}) or {}).get("name") or ""
        name_u = str(name).upper().strip()
        if name_u in valid_upper:
            label = valid_upper[name_u]
            out[f"HORAS_{label}"] += float(ch.get("hours") or 0)

    return {k: round(v, 2) for k, v in out.items()}


def aplicar_prioridades_horas_extra(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    franco  = pd.to_numeric(out["HORAS_FRANCO"],        errors="coerce").fillna(0.0)
    ex100   = pd.to_numeric(out["HORAS_EXTRA AL 100"],  errors="coerce").fillna(0.0)
    feriado = pd.to_numeric(out["HORAS_FERIADO"],       errors="coerce").fillna(0.0)
    ex50    = pd.to_numeric(out["HORAS_EXTRA AL 50"],   errors="coerce").fillna(0.0)

    es_noche = out["Turno"].str.strip().str.lower() == "noche"
    es_domingo_noche = es_noche & (out["dia"].str.strip().str.upper() == "DOMINGO")
    es_sabado_noche  = es_noche & (out["dia"].str.strip().str.upper() == "SÁBADO")

    # Flag de Regla A
    regla_a = out["_regla_a_aplicada"].fillna(False).astype(bool) \
        if "_regla_a_aplicada" in out.columns \
        else pd.Series(False, index=out.index)

    # ── EXTRA AL 100 ──
    # Para turno noche en domingo/sábado: solo se genera AL100 si es franco EN DÍA LIBRE
    # (planif=0). Si tiene planif>0 aunque la API diga FRANCO, no es un franco real.
    planif = pd.to_numeric(
        out["Horas planificadas"] if "Horas planificadas" in out.columns else 0,
        errors="coerce"
    ).fillna(0.0)
    eps_p = 1e-9
    franco_noche_libre = (franco > 0) & (planif < eps_p) & (es_domingo_noche | es_sabado_noche)

    out["HORAS_EXTRA AL 100"] = np.where(
        (feriado > 0) | regla_a,
        0.0,
        np.where(
            franco_noche_libre,
            franco,           # franco en día libre para noche → AL100
            np.where(
                es_domingo_noche | es_sabado_noche,
                0.0,          # noche sábado/domingo con planif → 0
                np.where(franco > 0, franco, ex100)  # resto: franco o ex100 de API
            )
        )
    ).round(2)

    # ── EXTRA AL 50 ──
    # AL50 solo aplica lunes-viernes y sábados hasta 13:00 → domingo y sábado noche no aplica
    out["HORAS_EXTRA AL 50"] = np.where(
        (feriado > 0) | regla_a | es_domingo_noche | es_sabado_noche,
        0.0,
        ex50
    ).round(2)

    return out


def corregir_horas_trabajadas(df: pd.DataFrame, tolerancia_horas: float = 1.0) -> pd.DataFrame:
    """
    Corrige Horas Trabajadas cuando la API reporta un valor inconsistente con el span
    real de fichadas (primera entrada → última salida), con una tolerancia de 1h para
    no afectar diferencias normales (redondeos, breaks, etc.).

    Casos cubiertos:
    - API sub-reporta (ej: Turco, overtime no computado): usa span de fichadas
    - API reporta 0 con fichadas válidas (ej: Cian en día libre): usa span de fichadas

    NO corrige fichajes claramente erróneos (ej: Barrera/Lemos con salida al día siguiente):
    esos deben corregirse en el origen (en Humand) para que queden visibles.
    """
    out = df.copy()

    worked_api = pd.to_numeric(out["_worked_api"], errors="coerce").fillna(0.0)
    has_fichadas = (~pd.isna(out["_rs"])) & (~pd.isna(out["_re"]))

    span = pd.Series(0.0, index=out.index)
    if has_fichadas.any():
        span[has_fichadas] = out.loc[has_fichadas].apply(
            lambda r: round(max(0.0, (r["_re"] - r["_rs"]).total_seconds() / 3600.0), 2),
            axis=1
        )

    use_span = has_fichadas & (np.abs(worked_api - span) > tolerancia_horas)
    out["Horas Trabajadas"] = np.where(use_span, span, worked_api).round(2)
    return out


def inferir_extras_faltantes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Cuando la API no categorizó horas extra pero el empleado trabajó más de lo
    planificado (o trabajó un día libre), infiere la categoría y el monto correctos.

    También setea HORAS_EXTRA (genérica) para que restar_llegada_anticipada funcione bien.

    Casos cubiertos:
    - Día planificado con overtime que la API no categorizó (ej: Turco lunes-viernes)
    - Día libre trabajado sin categorización de la API (ej: Cian domingo sin horario)
    """
    out = df.copy()
    eps = 1e-9

    def num(col):
        if col not in out.columns:
            return pd.Series(0.0, index=out.index)
        return pd.to_numeric(out[col], errors="coerce").fillna(0.0)

    worked  = num("Horas Trabajadas")
    planif  = num("Horas planificadas")
    feriado = num("HORAS_FERIADO")
    franco  = num("HORAS_FRANCO")
    ex50    = num("HORAS_EXTRA AL 50")
    ex100   = num("HORAS_EXTRA AL 100")
    extra_g = num("HORAS_EXTRA")

    sin_extras = (ex50 < eps) & (ex100 < eps) & (extra_g < eps) & (feriado < eps) & (franco < eps)

    dia_upper = out["dia"].str.strip().str.upper()
    es_noche  = out["Turno"].str.strip().str.lower() == "noche"
    es_domingo       = dia_upper == "DOMINGO"
    es_lunes_viernes = dia_upper.isin(["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"])

    def _media_hora(s):
        return s.apply(lambda x: math.floor(x / 0.5) * 0.5 if x >= 0.5 else 0.0).round(2)

    # ── Caso 1: día planificado con overtime no categorizado por la API ──
    hay_overtime = sin_extras & (planif > eps) & (worked > planif + 0.5)
    exceso_raw   = (worked - planif).clip(lower=0.0).round(2)

    mask_al50_ot  = hay_overtime & es_lunes_viernes
    mask_al100_ot = hay_overtime & es_domingo & (~es_noche)

    out.loc[mask_al50_ot,  "HORAS_EXTRA AL 50"]  = _media_hora(exceso_raw)[mask_al50_ot]  # AL50 redondeado
    out.loc[mask_al100_ot, "HORAS_EXTRA AL 100"] = exceso_raw[mask_al100_ot]               # AL100 sin redondear
    inferido_ot = mask_al50_ot | mask_al100_ot
    out.loc[inferido_ot, "HORAS_EXTRA"] = exceso_raw[inferido_ot]

    # ── Caso 2: planif=0 y trabajó → es franco ──
    # La API a veces no devuelve HORAS_FRANCO para días libres trabajados (ej: francos
    # entre semana). Se infiere franco y se limpia cualquier AL50 mal asignado por la API.
    es_franco_inferido = (planif < eps) & (worked > 0.5) & (franco < eps)

    out.loc[es_franco_inferido, "HORAS_FRANCO"]      = worked.round(2)[es_franco_inferido]
    out.loc[es_franco_inferido, "HORAS_EXTRA AL 50"] = 0.0

    return out